import numpy as np
from scipy import integrate
from scipy import optimize
import matplotlib.pyplot as plt
from openpyxl import Workbook

#extracting coordinates
loc = input("\nLocation of dataset: ")
m = float(input("Mass of particle (in kg): "))
g = 9.81
ds = np.loadtxt(loc)

t_list = ds[:, 0]
x_list = ds[:, 1]
y_list = ds[:, 2]

size = len(t_list)

print("\nComputation Initialized...")

#computing velocities and accelerations at given points
vx_list = []
vy_list = []
ax_list = []
ay_list = []

#forward difference - velocity
dt = t_list[1] - t_list[0]
vx_list.append((x_list[1] - x_list[0]) / dt)
vy_list.append((y_list[1] - y_list[0]) / dt)

#central differences - velocity
for i in range(1, size - 1):
    dt = t_list[i + 1] - t_list[i - 1]
    vx_list.append((x_list[i + 1] - x_list[i - 1]) / dt)
    vy_list.append((y_list[i + 1] - y_list[i - 1]) / dt)

#backward difference - velocity
dt = t_list[-1] - t_list[-2]
vx_list.append(((x_list[-1] - x_list[-2]) / dt))
vy_list.append(((y_list[-1] - y_list[-2]) / dt))

#forward difference - acceleration
dt = t_list[1] - t_list[0]
ax_list.append((vx_list[1] - vx_list[0]) / dt)
ay_list.append((vy_list[1] - vy_list[0]) / dt)

#central differences - acceleration
for i in range(1, size - 1):
    dt = t_list[i + 1] - t_list[i - 1]
    ax_list.append((vx_list[i + 1] - vx_list[i - 1]) / dt)
    ay_list.append((vy_list[i + 1] - vy_list[i - 1]) / dt)

#backward difference - acceleration
dt = t_list[-1] - t_list[-2]
ax_list.append((vx_list[-1] - vx_list[-2]) / dt)
ay_list.append((vy_list[-1] - vy_list[-2]) / dt)

vx_list = np.array(vx_list)
vy_list = np.array(vy_list)
ax_list = np.array(ax_list)
ay_list = np.array(ay_list)

#initial parameters - position and velocities
x0 = x_list[0]
y0 = y_list[0]
vx0 = vx_list[0]
vy0 = vy_list[0]

#piecewise hermite interpolation
def hermite(t1, t2, x1, x2, y1, y2,
           vx1, vx2, vy1, vy2,
           ax1, ax2, ay1, ay2):

    X = np.array([
        [1, t1, t1**2, t1**3, t1**4, t1**5],
        [1, t2, t2**2, t2**3, t2**4, t2**5],
        [0, 1, 2*t1, 3*t1**2, 4*t1**3, 5*t1**4],
        [0, 1, 2*t2, 3*t2**2, 4*t2**3, 5*t2**4],
        [0, 0, 2, 6*t1, 12*t1**2, 20*t1**3],
        [0, 0, 2, 6*t2, 12*t2**2, 20*t2**3]
    ], dtype=float)

    Y_forx = np.array([x1, x2, vx1, vx2, ax1, ax2], dtype=float)
    Y_fory = np.array([y1, y2, vy1, vy2, ay1, ay2], dtype=float)

    A_forx = np.linalg.solve(X, Y_forx)
    A_fory = np.linalg.solve(X, Y_fory)

    t_in = np.linspace(t1, t2, 100)

    x_in = np.zeros_like(t_in)
    y_in = np.zeros_like(t_in)

    for i in range(6):
        x_in += A_forx[i] * t_in**i
        y_in += A_fory[i] * t_in**i

    return t_in, x_in, y_in

#implementing interpolation over entire dataset
x_interp_all = []
y_interp_all = []
t_interp_all = []
for i in range(size - 1):
    t_seg, x_seg, y_seg = hermite(
        t_list[i], t_list[i+1],
        x_list[i], x_list[i+1],
        y_list[i], y_list[i+1],
        vx_list[i], vx_list[i+1],
        vy_list[i], vy_list[i+1],
        ax_list[i], ax_list[i+1],
        ay_list[i], ay_list[i+1]
    )

    if i != size - 2:
        t_seg = t_seg[:-1]
        x_seg = x_seg[:-1]
        y_seg = y_seg[:-1]

    t_interp_all.extend(t_seg)
    x_interp_all.extend(x_seg)
    y_interp_all.extend(y_seg)

t_interp_all = np.array(t_interp_all)
x_interp_all = np.array(x_interp_all)
y_interp_all = np.array(y_interp_all)

print("\nStage 1/3 Complete - Experimental trajectory processed.")

#equations of the projectile
vz0 = 0
def proj(t, state, b, k, wx, wy, wz, m):
    x, y, vx, vy, vz = state

    vrelx = vx-wx
    vrely = vy-wy
    vrelz = -wz
    vrel_mag = max(np.sqrt(vrelx**2 + vrely**2 + vrelz**2), 1e-12)

    dxdt = vx
    dydt = vy
    dvxdt = -(b/m)*(vrel_mag**(k-1))*vrelx
    dvydt = -g-((b/m)*(vrel_mag**(k-1))*vrely)
    dvzdt = -(b/m)*(vrel_mag**(k-1))*vrelz

    return dxdt, dydt, dvxdt, dvydt, dvzdt

#solving coupled ordinary differential equations
def objective(guess):
    b, k, wx, wy, wz = guess

    solution = integrate.solve_ivp(
    proj,
    (t_list[0], t_list[-1]),
    [x0, y0, vx0, vy0, vz0],
    t_eval=t_list,
    args=(b, k, wx, wy, wz, m),
    method="DOP853",
    rtol=1e-10,
    atol=1e-12
)

    if not solution.success:
        return np.inf

    x_fit = solution.y[0]
    y_fit = solution.y[1]

    error = np.sqrt(
        np.mean((x_list - x_fit)**2 + (y_list - y_fit)**2)
    )

    return error

#optimize for best fit values of b, k, wx, wy, wz
initial_guess = [0.05, 2, 0, 0, 0]

result = optimize.minimize(
    objective,
    initial_guess,
    method="L-BFGS-B",
    bounds=[
        (0, 10),
        (0, 5),
        (-120, 120),
        (-120, 120),
        (-120, 120)
    ]
)

if not result.success:
    raise RuntimeError(result.message)

b_fit, k_fit, wx_fit, wy_fit, wz_fit = result.x

print("\nStage 2/3 Complete - Aerodynamic parameters estimated.")

solution = integrate.solve_ivp(
    proj,
    (t_list[0], t_list[-1]),
    [x0, y0, vx0, vy0, vz0],
    t_eval=t_interp_all,
    args=(b_fit, k_fit, wx_fit, wy_fit, wz_fit, m),
    method="DOP853",
    rtol = 1e-10,
    atol = 1e-12
)

if not solution.success:
    raise RuntimeError(solution.message)

x_fit_dense = solution.y[0]
y_fit_dense = solution.y[1]
vx_fit_dense = solution.y[2]
vy_fit_dense = solution.y[3]
vz_fit_dense = solution.y[4]

ax_fit_dense = -(b_fit/m)*(np.sqrt((vx_fit_dense-wx_fit)**2 +
                                   (vy_fit_dense-wy_fit)**2 +
                                   (-wz_fit)**2)**(k_fit-1))*(vx_fit_dense-wx_fit)

ay_fit_dense = -g -(b_fit/m)*(np.sqrt((vx_fit_dense-wx_fit)**2 +
                                     (vy_fit_dense-wy_fit)**2 +
                                     (-wz_fit)**2)**(k_fit-1))*(vy_fit_dense-wy_fit)

az_fit_dense = -(b_fit/m)*(np.sqrt((vx_fit_dense-wx_fit)**2 +
                                   (vy_fit_dense-wy_fit)**2 +
                                   (-wz_fit)**2)**(k_fit-1))*(-wz_fit)

#trajectory length
trajectory_length = np.sum(np.hypot(np.diff(x_interp_all), np.diff(y_interp_all)))

percent_nrmsd = (np.sqrt(np.mean((x_interp_all - x_fit_dense)**2 + (y_interp_all - y_fit_dense)**2)) / trajectory_length) * 100

print("\nStage 3/3 Complete - Best-fit trajectory generated and model validated.")

#saving results
wb = Workbook()
ws = wb.active
ws.title = "Interpolated Trajectory Data"

# Trajectory data header
ws.append([
    "t [s]",
    "x [m]",
    "y [m]",
    "z [m]",
    "vx [m/s]",
    "vy [m/s]",
    "vz [m/s]",
    "ax [m/s^2]",
    "ay [m/s^2]",
    "az [m/s^2]"
])

# Trajectory data
for i in range(len(t_interp_all)):
    ws.append([
        t_interp_all[i],
        x_fit_dense[i],
        y_fit_dense[i],
        "-",
        vx_fit_dense[i],
        vy_fit_dense[i],
        "-",
        ax_fit_dense[i],
        ay_fit_dense[i],
        "-"
    ])

# Estimated parameters in columns L and M
ws["L1"] = "Estimated Model Parameters"
ws["M1"] = "Value"

ws["L2"] = "Drag coefficient (b)"
ws["M2"] = b_fit

ws["L3"] = "Drag exponent (k)"
ws["M3"] = k_fit

ws["L4"] = "Wind velocity (wx) [m/s]"
ws["M4"] = wx_fit

ws["L5"] = "Wind velocity (wy) [m/s]"
ws["M5"] = wy_fit

ws["L6"] = "Wind velocity (wz) [m/s]"
ws["M6"] = wz_fit

ws["L7"] = "Wind speed [m/s]"
ws["M7"] = np.sqrt(wx_fit**2 + wy_fit**2 + wz_fit**2)

ws["L8"] = "Trajectory length [m]"
ws["M8"] = trajectory_length

ws["L9"] = "% Normalized RMS deviation "
ws["M9"] = percent_nrmsd

wb.save("interpolated_trajectory_data.xlsx")
print("\nData Saved Succesfully.")

#reporting results
print("\nAssuming drag force = -b*(|v_rel|^k)v̂_rel:")

if b_fit <= 1e-6:
    print(f"\nb = {b_fit:.6f}")
    print("\nNo measurable drag was detected")
    print("Therefore, the drag exponent (k) and wind velocity are not identifiable from this dataset.")
else:
    print(f"b = {b_fit:.6f}")
    print(f"k = {k_fit:.6f}")
    print(f"Wind Velocity in x = {wx_fit:.4f} m/s")
    print(f"Wind Velocity in y = {wy_fit:.4f} m/s")
    print(f"Wind Velocity in z = {wz_fit:.4f} m/s")
    print(f"\nWind Speed: {np.sqrt(wx_fit**2 + wy_fit**2 + wz_fit**2):.4f} m/s")

print(f"\nNormalized RMS Deviation of Best Fit Model = {percent_nrmsd:.4f} %")
print(f"Trajectory Length = {trajectory_length:.2f} m")

view_plot = input("\nPress 'Enter' to view trajectory plots and deviation plot for the Best Fit Model and Interpolated Experimentally Observed data... ")

#plots
plt.figure(1)

#plt.plot(x_list, y_list, label="experimentally_observed_trajectory", color="green")
plt.plot(x_fit_dense, y_fit_dense, label="best_fit_trajectory", color="red")
plt.plot(x_interp_all, y_interp_all, label="interpolated_expt_trajectory", color="blue")

plt.xlabel("x [m]")
plt.ylabel("y [m]")
plt.title("Projectile Trajectories")
plt.legend()
plt.grid(True)

#plotting deviation
plt.figure(2)

deviation = np.hypot(
    x_interp_all - x_fit_dense,
    y_interp_all - y_fit_dense
)

plt.plot(
    t_interp_all,
    deviation,
    color="black"
)

plt.xlabel("time [s]")
plt.ylabel("Deviation [m]")
plt.title("Normalized RMS Deviation: Fitted Model vs. Interpolated Experimental Trajectory\n")
plt.grid(True)

#showing the plots
plt.show()

print("\n*Program has been terminated succesfully.\n")